# ===================================================================
# 1. IMPORTS & ENVIRONMENT SETUP
# ===================================================================
import os
import json
import re
import csv
import numpy as np
import asyncio
from datetime import datetime
from fastapi import FastAPI, HTTPException, BackgroundTasks
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from dotenv import load_dotenv

from langchain_core.prompts import ChatPromptTemplate
from langchain_groq import ChatGroq
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
import openpyxl

from medical_client import fetch_medical_dossier, send_notification
from sentiment_loader import get_sentiment_pipeline

load_dotenv()

class FeedbackRequest(BaseModel):
    user_id: int
    adherence_score: int  # 1-5
    recommendation_id: str  # optional, to track which herb recommendation

FEEDBACK_FILE = "adherence_feedback.csv"

# Create CSV file with headers if not exists
if not os.path.exists(FEEDBACK_FILE):
    with open(FEEDBACK_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "age", "bmi", "adherence_score", "success", "timestamp"])


async def store_feedback(fb: FeedbackRequest):
    """Store user's adherence score (1-5) after they tried the recommendation."""
    # You would normally retrieve the patient's age and BMI from your DB again
    # For simplicity, we'll fetch from the medical dossier
    try:
        patient = await fetch_medical_dossier(fb.user_id)
        age = patient.get('age', 40)
        bmi = patient.get('imc', 22)
        # Success is unknown at this point; we'll later ask "Did you reach your goal?"
        # For now, set success = None or 0
        with open(FEEDBACK_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([fb.user_id, age, bmi, fb.adherence_score, "", datetime.now().isoformat()])
        return {"status": "feedback recorded"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ===================================================================
# 2. GROQ LLM INITIALISATION
# ===================================================================
llm = ChatGroq(
    model="llama-3.1-8b-instant",
    temperature=0.3,
    api_key=os.getenv("GROQ_API_KEY")
)

# ===================================================================
# 3. MACHINE LEARNING OUTCOME PREDICTOR
# ===================================================================
class OutcomePredictor:
    def __init__(self):
        self.model = None
        self.scaler = StandardScaler()
        self.is_trained = False
        self._load_or_train()

    def _generate_synthetic_data(self):
        """Replace with real historical data from your database."""
        np.random.seed(42)
        ages = np.random.randint(18, 80, 200)
        bmis = np.random.uniform(15, 40, 200)
        adherence = np.random.randint(1, 6, 200)
        success = ((bmis >= 18.5) & (bmis < 25) & (adherence >= 3)).astype(int)
        noise = np.random.choice([0, 1], size=200, p=[0.95, 0.05])
        success = success | noise
        X = np.column_stack([ages, bmis, adherence])
        y = success
        return X, y

    def _load_or_train(self):
        X, y = self._generate_synthetic_data()
        X_scaled = self.scaler.fit_transform(X)
        self.model = LogisticRegression(random_state=42)
        self.model.fit(X_scaled, y)
        self.is_trained = True
        print("✅ Outcome predictor trained on synthetic data.")

    def predict_probability(self, age: float, bmi: float, adherence_score: float = 3.0) -> float:
        if not self.is_trained:
            self._load_or_train()
        features = np.array([[age, bmi, adherence_score]])
        features_scaled = self.scaler.transform(features)
        prob = self.model.predict_proba(features_scaled)[0][1]
        return float(prob)

outcome_predictor = OutcomePredictor()

# ===================================================================
# 4. DYNAMIC TCM LOOKUP FROM EXCEL FILES
# ===================================================================
EXCEL_PATH = "data"
_herb_cache = {}
_disease_cache = {}
_ingredient_cache = {}
_all_herb_names = []

def _load_workbook_readonly(filename: str):
    full_path = os.path.join(EXCEL_PATH, filename)
    if not os.path.exists(full_path):
        print(f"⚠️ Warning: {full_path} not found")
        return None
    return openpyxl.load_workbook(full_path, read_only=True)

def _get_header_map(sheet):
    headers = {}
    for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx
    return headers

def lookup_herb_by_name(herb_name: str) -> Optional[Dict[str, Any]]:
    if not herb_name:
        return None
    key = herb_name.lower()
    if key in _herb_cache:
        return _herb_cache[key]

    wb = _load_workbook_readonly("herb_all.xlsx")
    if not wb:
        return None
    sheet = wb.active
    header_map = _get_header_map(sheet)

    name_col = header_map.get('herb_name') or header_map.get('name') or header_map.get('herb')
    indications_col = header_map.get('indications') or header_map.get('action')
    contraindications_col = header_map.get('contraindications') or header_map.get('contraindication')
    ingredients_col = header_map.get('ingredients') or header_map.get('chemical composition')

    if not name_col:
        wb.close()
        return None

    result = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[name_col-1] and herb_name.lower() in str(row[name_col-1]).lower():
            result = {
                "name": row[name_col-1],
                "indications": row[indications_col-1] if indications_col and indications_col-1 < len(row) else "",
                "contraindications": row[contraindications_col-1] if contraindications_col and contraindications_col-1 < len(row) else "",
                "ingredients": row[ingredients_col-1] if ingredients_col and ingredients_col-1 < len(row) else ""
            }
            break
    wb.close()
    _herb_cache[key] = result
    return result

def lookup_herbs_by_disease(disease_name: str) -> List[str]:
    if not disease_name:
        return []
    key = disease_name.lower()
    if key in _disease_cache:
        return _disease_cache[key]

    wb = _load_workbook_readonly("disease_all.xlsx")
    if not wb:
        return []
    sheet = wb.active
    header_map = _get_header_map(sheet)
    disease_col = header_map.get('disease') or header_map.get('disease_name')
    herb_col = header_map.get('herb') or header_map.get('herb_name')
    if not disease_col or not herb_col:
        wb.close()
        return []

    herbs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[disease_col-1] and disease_name.lower() in str(row[disease_col-1]).lower():
            herb = row[herb_col-1]
            if herb and herb not in herbs:
                herbs.append(str(herb))
    wb.close()
    _disease_cache[key] = herbs
    return herbs

def lookup_herbs_by_ingredient(ingredient_name: str) -> List[str]:
    if not ingredient_name:
        return []
    key = ingredient_name.lower()
    if key in _ingredient_cache:
        return _ingredient_cache[key]

    wb = _load_workbook_readonly("ingredient_all.xlsx")
    if not wb:
        return []
    sheet = wb.active
    header_map = _get_header_map(sheet)
    ing_col = header_map.get('ingredient') or header_map.get('compound')
    herb_col = header_map.get('herb') or header_map.get('herb_name')
    if not ing_col or not herb_col:
        wb.close()
        return []

    herbs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[ing_col-1] and ingredient_name.lower() in str(row[ing_col-1]).lower():
            herb = row[herb_col-1]
            if herb and herb not in herbs:
                herbs.append(str(herb))
    wb.close()
    _ingredient_cache[key] = herbs
    return herbs

def find_tcm_herbs_dynamic(user_message: str, patient_diseases: List[str] = None) -> List[Dict[str, Any]]:
    if patient_diseases is None:
        patient_diseases = []

    user_lower = user_message.lower()
    candidate_herb_names = set()

    # 1. Disease matching
    common_diseases = ["cold", "fatigue", "insomnia", "anxiety", "indigestion", "diabetes", "hypertension", "cough"]
    for disease in common_diseases:
        if disease in user_lower:
            herbs = lookup_herbs_by_disease(disease)
            candidate_herb_names.update(herbs)

    # 2. Ingredient matching
    common_ingredients = ["berberine", "quercetin", "ginsenoside", "glycyrrhizin", "baicalin"]
    for ing in common_ingredients:
        if ing in user_lower:
            herbs = lookup_herbs_by_ingredient(ing)
            candidate_herb_names.update(herbs)

    # 3. Direct herb name extraction
    global _all_herb_names
    if not _all_herb_names:
        wb = _load_workbook_readonly("herb_all.xlsx")
        if wb:
            sheet = wb.active
            header_map = _get_header_map(sheet)
            name_col = header_map.get('herb_name') or header_map.get('name')
            if name_col:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[name_col-1]:
                        _all_herb_names.append(str(row[name_col-1]).lower())
            wb.close()
    for herb_name_lower in _all_herb_names[:1000]:
        if herb_name_lower in user_lower:
            candidate_herb_names.add(herb_name_lower.title())

    # 4. Fetch details and filter contraindications
    suggestions = []
    for herb_name in candidate_herb_names:
        herb_info = lookup_herb_by_name(herb_name)
        if not herb_info:
            continue
        contra_text = herb_info.get("contraindications", "")
        if contra_text and patient_diseases:
            contra_lower = contra_text.lower()
            if any(d.lower() in contra_lower for d in patient_diseases):
                continue
        suggestions.append(herb_info)
        if len(suggestions) >= 3:
            break

    # 5. Fallback: keyword search on indications
    if not suggestions:
        wb = _load_workbook_readonly("herb_all.xlsx")
        if wb:
            sheet = wb.active
            header_map = _get_header_map(sheet)
            name_col = header_map.get('herb_name') or header_map.get('name')
            ind_col = header_map.get('indications') or header_map.get('action')
            contra_col = header_map.get('contraindications')
            if name_col and ind_col:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    herb_name = row[name_col-1]
                    indications = row[ind_col-1] or ""
                    if any(word in user_lower for word in indications.lower().split()[:8]):
                        if contra_col and patient_diseases:
                            contra = row[contra_col-1] or ""
                            if any(d.lower() in contra.lower() for d in patient_diseases):
                                continue
                        suggestions.append({
                            "name": herb_name,
                            "indications": indications,
                            "contraindications": row[contra_col-1] if contra_col else "",
                            "ingredients": ""
                        })
                        if len(suggestions) >= 3:
                            break
            wb.close()
    return suggestions

# ===================================================================
# 5. HELPER FUNCTION TO EXTRACT JSON FROM LLM RESPONSE
# ===================================================================
def extract_json_from_response(content: str) -> dict:
    """Safely extract JSON from LLM response that may contain extra text."""
    json_match = re.search(r'\{.*\}', content, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group())
        except json.JSONDecodeError:
            pass
    return {"subject": "Gentle reminder", "body": "Take care of yourself today."}

# ===================================================================
# 6. IMMEDIATE CARING MESSAGE (with escaped braces)
# ===================================================================
async def generate_care_message(patient: dict, user_message: str, sentiment_score: float) -> dict:
    care_prompt = ChatPromptTemplate.from_messages([
        ("system", """You are a compassionate health assistant. 
The user is feeling down (sentiment score: {sentiment_score}). 
Generate a JSON object with two fields: "subject" and "body".
Example: {{"subject": "You are not alone", "body": "Take a deep breath. Herbal tea might help."}}
Do not include any other text.

Use the patient context only to make the message relevant, but keep it light and supportive.
Do NOT give medical advice. Stay within 2‑3 sentences.

Patient context:
- Diseases: {diseases}
- Allergies: {allergies}
- Current treatments: {treatments}
- Recent user message: "{user_message}"
"""),
        ("human", "Generate the JSON.")
    ])
    chain = care_prompt | llm
    response = await chain.ainvoke({
        "sentiment_score": sentiment_score,
        "diseases": patient.get("maladies", "none"),
        "allergies": patient.get("allergies", "none"),
        "treatments": patient.get("traitements", "none"),
        "user_message": user_message[:200]
    })
    try:
        content = response.content.strip()
        data = extract_json_from_response(content)
        subject = data.get("subject", "Gentle reminder")
        body = data.get("body", "Take care of yourself today.")
    except Exception:
        subject = "Your well-being matters"
        body = "We noticed you might be feeling down. Gentle TCM herbs like chamomile or peppermint can help you relax. Take care."
    return {"subject": subject, "body": body}

# ===================================================================
# 7. DELAYED FOLLOW-UP EMAIL (1 minute later)
# ===================================================================
async def send_delayed_followup_email(
    email: str,
    user_message: str,
    patient: dict,
    tcm_herbs: List[Dict[str, Any]],
    sentiment_label: str,
    sentiment_score: float
):
    await asyncio.sleep(60)  # 1 minutes

    herbs_text = "\n".join([f"- {h['name']}: {h.get('indications', 'No details')}" for h in tcm_herbs])
    if not herbs_text:
        herbs_text = "No specific herbs were suggested, but general wellness practices apply."

    followup_prompt = ChatPromptTemplate.from_messages([
        ("system", """You are a compassionate health coach. Generate a **complete email** (plain text, not JSON) that includes:

1. **How to use the suggested TCM herbs** – explain safe preparation, dosage, and timing. Use the herbs already recommended.
2. **Personalised sports routine** – suggest simple physical activities based on the patient's mental state ({sentiment_label}) and diseases ({diseases}).
3. **A motivational quote** – short, uplifting, relevant.
4. **Positive reminder** – encourage the user that they can achieve their health goals.

The email should feel warm, supportive, and dynamic – no generic templates.
Write the email directly, starting with "Subject: Your personalised follow-up". Keep total under 500 words.

Patient info:
- Diseases: {diseases}
- Allergies: {allergies}
- Sentiment: {sentiment_label} (score {sentiment_score:.2f})

User's original message: "{user_message}"

Suggested TCM herbs:
{tcm_herbs_text}
"""),
        ("human", "Generate the follow-up email.")
    ])

    chain = followup_prompt | llm
    response = await chain.ainvoke({
        "sentiment_label": sentiment_label,
        "sentiment_score": sentiment_score,
        "diseases": patient.get("maladies", "none"),
        "allergies": patient.get("allergies", "none"),
        "user_message": user_message[:300],
        "tcm_herbs_text": herbs_text
    })

    email_content = response.content
    lines = email_content.split("\n", 1)
    subject = lines[0].replace("Subject:", "").strip() if lines[0].startswith("Subject") else "Your personalised health follow-up"
    body = lines[1] if len(lines) > 1 else email_content

    await send_notification("email", email, subject=subject, body=body)
    print(f"📧 Delayed follow-up email sent to {email}")

# ===================================================================
# 8. FASTAPI APP, REQUEST/RESPONSE MODELS, AND MAIN ENDPOINT
# ===================================================================
app = FastAPI(title="Health AI Composer")

# ----- Feedback endpoints (store adherence scores) -----
@app.post("/feedback")
async def store_feedback(fb: FeedbackRequest):
    """Store user's adherence score (1-5) after they tried the recommendation."""
    try:
        patient = await fetch_medical_dossier(fb.user_id)
        age = patient.get('age', 40)
        bmi = patient.get('imc', 22)
        with open(FEEDBACK_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([fb.user_id, age, bmi, fb.adherence_score, "", datetime.now().isoformat()])
        return {"status": "feedback recorded"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/feedback")
async def get_feedback(user_id: int):
    """Optional: retrieve past feedback scores (for testing)."""
    # Simple implementation: read CSV and filter by user_id
    feedbacks = []
    if os.path.exists(FEEDBACK_FILE):
        with open(FEEDBACK_FILE, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if int(row['user_id']) == user_id:
                    feedbacks.append(row)
    return {"user_id": user_id, "feedbacks": feedbacks}

# ----- Chat request/response models -----
class ChatRequest(BaseModel):
    user_id: int
    message: str
    email: Optional[str] = None
    phone: Optional[str] = None

class ChatResponse(BaseModel):
    response: str
    sentiment: str
    tcm_herbs_suggested: List[str]
    predicted_outcome: str


# Main consultation prompt (escaped braces)
main_prompt = ChatPromptTemplate.from_messages([
    ("system", """You are an AI health assistant specialised in clinical nutrition and Traditional Chinese Medicine (TCM).

Patient data:
- Weight: {weight} kg
- Height: {height} cm
- BMI: {bmi}
- Allergies: {allergies} (severity: {allergy_severity})
- Diseases: {diseases}
- Current treatments: {treatments}

TCM herbal suggestions based on patient's symptoms:
{tcm_suggestions}

Machine learning outcome prediction: There is a {success_probability}% chance of achieving health goals with consistent adherence.

Rules:
- Never recommend foods or herbs that conflict with allergies or medications.
- Always add: "Consult a licensed TCM practitioner before using any herbs."
- Give a realistic, conservative outcome prediction.
- Keep the answer helpful, concise, and empathetic.
"""),
    ("human", "{question}")
])

@app.post("/chat", response_model=ChatResponse)
async def chat_compose(req: ChatRequest, background_tasks: BackgroundTasks):
    # 1. Fetch patient medical dossier
    try:
        patient = await fetch_medical_dossier(req.user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot retrieve medical data: {str(e)}")

    # 2. Sentiment analysis
    sentiment_pipe = get_sentiment_pipeline()
    sentiment = sentiment_pipe(req.message)[0]
    sentiment_label = sentiment['label']
    sentiment_score = sentiment['score']

    # 3. Dynamic TCM herb matching
    diseases = patient.get("maladies", "")
    disease_list = [d.strip() for d in diseases.split(",")] if diseases else []
    herb_matches = find_tcm_herbs_dynamic(req.message, disease_list)
    tcm_text = "\n".join([f"- {h['name']}: {h.get('indications', '')}" for h in herb_matches]) if herb_matches else "No specific herb identified."

    # 4. ML outcome prediction
    age = patient.get('age', 40)
    bmi = patient.get('imc')
    if bmi is None or bmi == '':
        bmi = 22.0
    else:
        bmi = float(bmi)
    adherence = 3.0
    success_prob = outcome_predictor.predict_probability(age, bmi, adherence)
    predicted_outcome = f"Based on machine learning analysis of similar patients, the estimated chance of reaching your health goals is {success_prob*100:.1f}%."

    # 5. Main AI response
    chain = main_prompt | llm
    response = await chain.ainvoke({
        "weight": patient.get("poids", "unknown"),
        "height": patient.get("taille", "unknown"),
        "bmi": patient.get("imc", "unknown"),
        "allergies": patient.get("allergies", "none"),
        "allergy_severity": patient.get("allergie_severity", "unknown"),
        "diseases": patient.get("maladies", "none"),
        "treatments": patient.get("traitements", "none"),
        "tcm_suggestions": tcm_text,
        "success_probability": round(success_prob * 100, 1),
        "question": req.message
    })

    # 6. Immediate caring message for very negative sentiment
    if sentiment_label == "NEGATIVE" and sentiment_score > 0.85:
        care_msg = await generate_care_message(patient, req.message, sentiment_score)
        if req.email:
            await send_notification("email", req.email,
                                    subject=care_msg["subject"],
                                    body=care_msg["body"])
        if req.phone:
            sms_body = care_msg["body"][:160]
            await send_notification("sms", req.phone, body=sms_body)

    # 7. Schedule delayed follow-up email (3 minutes)
    if req.email:
        background_tasks.add_task(
            send_delayed_followup_email,
            email=req.email,
            user_message=req.message,
            patient=patient,
            tcm_herbs=herb_matches,
            sentiment_label=sentiment_label,
            sentiment_score=sentiment_score
        )

    # 8. Return immediate response
    return ChatResponse(
        response=response.content,
        sentiment=sentiment_label,
        tcm_herbs_suggested=[h["name"] for h in herb_matches],
        predicted_outcome=predicted_outcome
    )

# ===================================================================
# 9. RUN SERVER
# ===================================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)